import os
import json
import argparse
from typing import Dict, List
from flask import Flask, render_template, request, jsonify, Response, stream_with_context, redirect, url_for, session
from functools import wraps
from dotenv import load_dotenv
from datetime import date, timedelta, datetime
import re
import pytz
from apscheduler.schedulers.background import BackgroundScheduler
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

from sheet_client import fetch_grouped_messages, load_settings, inspect_sheets, diagnose_matches, fetch_grouped_messages_by_date, stream_grouped_messages_by_date, mark_checked_for_agency, mark_checked_for_agencies, list_sheet_tabs, inspect_sheets_by_id, compute_settlement_rows
from internal_manager import load_cache as internal_load_cache, refresh_cache as internal_refresh_cache, fetch_workload_schedule
from guarantee_manager import GuaranteeManager
from auth import AuthManager, ROLES

try:
	from data_security import DataSecurity
	SECURITY_AVAILABLE = True
except ImportError:
	SECURITY_AVAILABLE = False


# .env 로드
load_dotenv()


def _strip_parentheses(text: str) -> str:
	"""문자열에서 소괄호 내 내용을 제거한다. 예: '작업명(부가)' -> '작업명'"""
	if not text:
		return text
	return re.sub(r"\s*\([^)]*\)", "", text).strip()


# 로그인 필수 데코레이터
def login_required(f):
	@wraps(f)
	def decorated_function(*args, **kwargs):
		if not session.get("user"):
			# API 요청인 경우 JSON 응답
			if request.path.startswith("/api/"):
				return jsonify({"error": "unauthorized", "message": "로그인이 필요합니다."}), 401
			# 페이지 요청인 경우 로그인 페이지로 리다이렉트
			return redirect(url_for("login_page"))
		return f(*args, **kwargs)
	return decorated_function


# 관리자 권한 필수 데코레이터
def admin_required(f):
	@wraps(f)
	def decorated_function(*args, **kwargs):
		user = session.get("user")
		if not user:
			if request.path.startswith("/api/"):
				return jsonify({"error": "unauthorized"}), 401
			return redirect(url_for("login_page"))
		if user.get("role") != "admin":
			if request.path.startswith("/api/"):
				return jsonify({"error": "forbidden", "message": "관리자 권한이 필요합니다."}), 403
			return redirect(url_for("index"))
		return f(*args, **kwargs)
	return decorated_function


def create_app() -> Flask:
	app = Flask(__name__)
	
	# 세션 암호화 키 설정
	app.secret_key = os.getenv("SECRET_KEY", "deadline-notifier-secret-key-change-me")
	app.config["SESSION_COOKIE_SECURE"] = os.getenv("FLASK_ENV") == "production"
	app.config["SESSION_COOKIE_HTTPONLY"] = True
	app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
	app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(days=7)
	
	# 인증 매니저 초기화
	auth_manager = AuthManager()
	
	# 스케줄러 초기화
	scheduler = BackgroundScheduler(timezone=pytz.timezone('Asia/Seoul'))
	scheduler.start()
	
	# Render 배포: 앱 시작 시 자동 초기화
	if os.getenv("AUTO_SYNC_ON_START", "false").lower() == "true":
		logger.info("🚀 Auto-sync on startup enabled (Render mode)")
		
		def init_on_startup():
			"""앱 시작 시 데이터 초기화"""
			import time
			time.sleep(5)  # 앱 완전 시작 대기
			
			try:
				logger.info("📡 Starting auto-sync from Google Sheets...")
				gm = GuaranteeManager()
				result = gm.sync_from_google_sheets()
				logger.info(f"✅ Auto-sync completed: {result}")
			except Exception as e:
				logger.error(f"❌ Auto-sync failed: {e}")
			
			try:
				logger.info("⚡ Starting workload cache refresh...")
				from workload_cache import refresh_all_workload_cache
				result = refresh_all_workload_cache()
				logger.info(f"✅ Workload cache refreshed: {result['message']}")
			except Exception as e:
				logger.error(f"❌ Workload cache refresh failed: {e}")
		
		# 백그라운드 스레드로 실행
		import threading
		init_thread = threading.Thread(target=init_on_startup)
		init_thread.daemon = True
		init_thread.start()
	
	# 자동 동기화 태스크
	def sync_guarantee_data():
		"""보장건 데이터 자동 동기화"""
		try:
			logger.info(f"Starting automatic sync at {datetime.now(pytz.timezone('Asia/Seoul'))}")
			gm = GuaranteeManager()
			result = gm.sync_from_google_sheets()
			logger.info(f"Sync completed: {result}")
		except Exception as e:
			logger.error(f"Sync failed: {e}")
	
	# 작업량 캐시 자동 갱신 태스크
	def refresh_workload_cache():
		"""작업량 캐시 자동 갱신"""
		try:
			logger.info(f"Starting workload cache refresh at {datetime.now(pytz.timezone('Asia/Seoul'))}")
			from workload_cache import refresh_all_workload_cache
			result = refresh_all_workload_cache()
			logger.info(f"Workload cache refresh completed: {result['message']}")
		except Exception as e:
			logger.error(f"Workload cache refresh failed: {e}")
	
	# 스케줄러 잠금 (동시 실행 방지)
	import threading
	_scheduler_lock = threading.Lock()
	_scheduler_running = {"rank_crawl": False}
	
	# 순위 크롤링 자동 실행 태스크 (N2 포함, Google Sheets 저장)
	def crawl_ranks_auto():
		"""순위 자동 크롤링 (N2 포함)
		
		주의: Render에서 workers=1 권장. 또는 외부 cron이 token endpoint를 호출하는 방식 사용.
		동시 실행 방지를 위해 잠금 사용.
		"""
		# 이미 실행 중인지 확인
		if _scheduler_running.get("rank_crawl"):
			logger.warning("⚠️ Rank crawling already running, skipping...")
			return
		
		with _scheduler_lock:
			_scheduler_running["rank_crawl"] = True
		
		from scheduler_logs import log_scheduler_event
		log_scheduler_event("rank_crawl", "순위 크롤링", "started", "크롤링 시작")
		try:
			logger.info(f"🏆 Starting automatic rank crawling at {datetime.now(pytz.timezone('Asia/Seoul'))}")
			from rank_crawler import crawl_ranks_for_company
			
			# 전체 회사 한 번에 크롤링 (None = 모두)
			result = crawl_ranks_for_company(None)
			logger.info(f"✅ Rank crawling completed: {result.get('message', 'Unknown')}")
			logger.info(f"   Crawled: {result.get('crawled_count', 0)}, Failed: {result.get('failed_count', 0)}")
			log_scheduler_event("rank_crawl", "순위 크롤링", "success", 
				f"크롤링 {result.get('crawled_count', 0)}건 완료", result)
			
			# 15시 크롤링인 경우 보장건 시트 자동 업데이트
			current_hour = datetime.now(pytz.timezone('Asia/Seoul')).hour
			if current_hour >= 12:  # 오후 크롤링인 경우
				# 보장건 시트 업데이트
				try:
					log_scheduler_event("guarantee_update", "보장건 시트 업데이트", "started", "업데이트 시작")
					logger.info("📋 Updating guarantee sheets...")
					from rank_update_service import update_guarantee_sheets_from_snapshots
					update_result = update_guarantee_sheets_from_snapshots()
					logger.info(f"✅ Guarantee sheets updated: {update_result}")
					log_scheduler_event("guarantee_update", "보장건 시트 업데이트", "success", "업데이트 완료")
				except Exception as update_error:
					logger.error(f"❌ Guarantee sheet update failed: {update_error}")
					log_scheduler_event("guarantee_update", "보장건 시트 업데이트", "failed", str(update_error))
				
				# 레시피 분석 실행
				try:
					log_scheduler_event("recipe_analysis", "레시피 분석", "started", "분석 시작")
					logger.info("📊 Running recipe analysis...")
					from recipe_analyzer import get_analyzer
					analyzer = get_analyzer()
					analysis_result = analyzer.analyze_all(weeks=3)
					logger.info(f"✅ Recipe analysis complete: {analysis_result.get('total_analyzed', 0)} businesses")
					log_scheduler_event("recipe_analysis", "레시피 분석", "success", 
						f"{analysis_result.get('total_analyzed', 0)}개 업체 분석 완료")
				except Exception as analysis_error:
					logger.error(f"❌ Recipe analysis failed: {analysis_error}")
					log_scheduler_event("recipe_analysis", "레시피 분석", "failed", str(analysis_error))
				
				# 학습 데이터셋 빌드 (크롤링 후 자동 실행)
				try:
					log_scheduler_event("training_build", "학습 데이터셋 빌드", "started", "빌드 시작")
					logger.info("🎓 Building training dataset...")
					from training_dataset_builder import build_and_save
					build_result = build_and_save(weeks=3)
					logger.info(f"✅ Training dataset built: {build_result.get('training_rows_count', 0)} rows")
					log_scheduler_event("training_build", "학습 데이터셋 빌드", "success", 
						f"{build_result.get('training_rows_count', 0)}행 생성 완료")
				except Exception as build_error:
					logger.error(f"❌ Training dataset build failed: {build_error}")
					log_scheduler_event("training_build", "학습 데이터셋 빌드", "failed", str(build_error))
					
		except Exception as e:
			logger.error(f"❌ Automatic rank crawling failed: {e}")
			import traceback
			logger.error(traceback.format_exc())
			log_scheduler_event("rank_crawl", "순위 크롤링", "failed", str(e))
		finally:
			# 잠금 해제
			with _scheduler_lock:
				_scheduler_running["rank_crawl"] = False
			logger.info("🔓 Rank crawling lock released")
	
	# 매일 9시, 16시 스케줄 등록 (보장건 동기화)
	scheduler.add_job(func=sync_guarantee_data, trigger="cron", hour=9, minute=0, id="morning_sync")
	scheduler.add_job(func=sync_guarantee_data, trigger="cron", hour=16, minute=0, id="afternoon_sync")
	
	# 매일 11:20 스케줄 등록 (Worklog 캐시 갱신)
	def refresh_worklog_cache_task():
		"""Worklog 캐시 자동 갱신"""
		from scheduler_logs import log_scheduler_event
		log_scheduler_event("worklog_cache", "Worklog 캐시", "started", "캐시 갱신 시작")
		try:
			logger.info(f"📝 Starting worklog cache refresh at {datetime.now(pytz.timezone('Asia/Seoul'))}")
			from worklog_cache import refresh_worklog_cache as _refresh_worklog
			result = _refresh_worklog()
			logger.info(f"✅ Worklog cache refresh completed: {result.get('message')}")
			log_scheduler_event("worklog_cache", "Worklog 캐시", "success", 
				f"{result.get('records_count', 0)}건 갱신 완료", result)
		except Exception as e:
			logger.error(f"❌ Worklog cache refresh failed: {e}")
			log_scheduler_event("worklog_cache", "Worklog 캐시", "failed", str(e))
	
	scheduler.add_job(func=refresh_worklog_cache_task, trigger="cron", hour=3, minute=30, id="worklog_cache_refresh_task")
	
	# 매일 03:00 스케줄 등록 (작업량 캐시 갱신 - 새벽 시간대로 변경)
	scheduler.add_job(func=refresh_workload_cache, trigger="cron", hour=3, minute=0, id="workload_cache_refresh")
	
	# 매일 15:10 스케줄 등록 (순위 크롤링 - 1일 1회)
	# 주의: Render workers=1이 아니면 외부 cron 사용 권장
	if os.getenv("USE_INTERNAL_SCHEDULER", "true").lower() == "true":
		scheduler.add_job(func=crawl_ranks_auto, trigger="cron", hour=15, minute=10, id="daily_rank_crawl")
		logger.info("📅 Internal scheduler enabled:")
		logger.info("   - 03:00 작업량 캐시 갱신")
		logger.info("   - 03:30 Worklog 캐시 갱신")
		logger.info("   - 15:10 순위 크롤링")
	else:
		logger.info("📅 Internal scheduler disabled. Use /api/cron/crawl-ranks with CRON_TOKEN")

	# --- 인증 관련 라우트 ---
	@app.route("/login", methods=["GET"])
	def login_page():
		"""로그인 페이지"""
		# 이미 로그인된 경우 메인으로 리다이렉트
		if session.get("user"):
			return redirect(url_for("index"))
		return render_template("login.html")
	
	@app.route("/api/auth/login", methods=["POST"])
	def api_auth_login():
		"""로그인 API"""
		try:
			data = request.get_json(force=True) or {}
		except Exception:
			return jsonify({"error": "invalid_json"}), 400
		
		username = str(data.get("username", "")).strip()
		password = str(data.get("password", ""))
		
		if not username or not password:
			return jsonify({"error": "아이디와 비밀번호를 입력해주세요."}), 400
		
		user = auth_manager.authenticate(username, password)
		if user:
			session.permanent = True
			session["user"] = user
			logger.info(f"User logged in: {username}")
			return jsonify({"ok": True, "user": user, "redirect": "/"}), 200
		
		return jsonify({"error": "아이디 또는 비밀번호가 일치하지 않습니다."}), 401
	
	@app.route("/api/auth/logout", methods=["POST"])
	def api_auth_logout():
		"""로그아웃 API"""
		user = session.get("user")
		if user:
			logger.info(f"User logged out: {user.get('username')}")
		session.clear()
		return jsonify({"ok": True}), 200
	
	@app.route("/api/auth/me", methods=["GET"])
	@login_required
	def api_auth_me():
		"""현재 로그인 사용자 정보"""
		return jsonify({"user": session.get("user")}), 200
	
	# --- 계정 관리 API (관리자 전용) ---
	@app.route("/api/admin/users", methods=["GET"])
	@admin_required
	def api_admin_users():
		"""사용자 목록 조회"""
		include_inactive = request.args.get("include_inactive", "").lower() == "true"
		users = auth_manager.get_all_users(include_inactive)
		return jsonify({"users": users, "count": len(users)}), 200
	
	@app.route("/api/admin/users", methods=["POST"])
	@admin_required
	def api_admin_create_user():
		"""사용자 생성"""
		try:
			data = request.get_json(force=True) or {}
		except Exception:
			return jsonify({"error": "invalid_json"}), 400
		
		username = str(data.get("username", "")).strip()
		password = str(data.get("password", ""))
		role = str(data.get("role", "user")).strip()
		name = str(data.get("name", "")).strip()
		
		if not username or not password:
			return jsonify({"error": "아이디와 비밀번호는 필수입니다."}), 400
		
		current_user = session.get("user", {})
		user = auth_manager.create_user(username, password, role, name, current_user.get("id"))
		
		if user:
			return jsonify({"ok": True, "user": user}), 201
		return jsonify({"error": "이미 존재하는 아이디입니다."}), 400
	
	@app.route("/api/admin/users/<user_id>", methods=["PUT"])
	@admin_required
	def api_admin_update_user(user_id):
		"""사용자 수정"""
		try:
			data = request.get_json(force=True) or {}
		except Exception:
			return jsonify({"error": "invalid_json"}), 400
		
		current_user = session.get("user", {})
		user = auth_manager.update_user(user_id, data, current_user.get("id"))
		
		if user:
			return jsonify({"ok": True, "user": user}), 200
		return jsonify({"error": "사용자를 찾을 수 없습니다."}), 404
	
	@app.route("/api/admin/users/<user_id>", methods=["DELETE"])
	@admin_required
	def api_admin_delete_user(user_id):
		"""사용자 삭제"""
		current_user = session.get("user", {})
		
		# 자기 자신은 삭제 불가
		if user_id == current_user.get("id"):
			return jsonify({"error": "자신의 계정은 삭제할 수 없습니다."}), 400
		
		if auth_manager.delete_user(user_id, current_user.get("id")):
			return jsonify({"ok": True}), 200
		return jsonify({"error": "삭제할 수 없습니다."}), 400
	
	@app.route("/api/admin/users/<user_id>/password", methods=["PUT"])
	@admin_required
	def api_admin_change_password(user_id):
		"""비밀번호 변경 (관리자용)"""
		try:
			data = request.get_json(force=True) or {}
		except Exception:
			return jsonify({"error": "invalid_json"}), 400
		
		new_password = str(data.get("password", ""))
		if not new_password:
			return jsonify({"error": "새 비밀번호를 입력해주세요."}), 400
		
		if auth_manager.change_password(user_id, new_password):
			return jsonify({"ok": True}), 200
		return jsonify({"error": "사용자를 찾을 수 없습니다."}), 404
	
	@app.route("/api/admin/roles", methods=["GET"])
	@login_required
	def api_admin_roles():
		"""역할 목록 조회"""
		return jsonify({"roles": ROLES}), 200

	@app.route("/", methods=["GET"])  # 메인 페이지: 폼 + 결과
	@login_required
	def index():
		settings = load_settings()
		days_param = request.args.get("days", "").strip()
		base_date_str = request.args.get("base_date", "").strip()
		filter_mode = request.args.get("filter_mode", "agency").strip().lower()  # 'agency' | 'internal'
		did_fetch = request.args.get("submit", "") == "1"

		# 기준일 처리 (기본: 오늘)
		if base_date_str:
			try:
				base_dt = date.fromisoformat(base_date_str)
			except ValueError:
				base_dt = date.today()
		else:
			base_dt = date.today()

		ordered_days: List[int] = []
		day_to_date: Dict[int, str] = {}
		day_to_date_label: Dict[int, str] = {}
		grouped_by_date: Dict[str, Dict[int, Dict[str, List[str]]]] = {}
		agency_to_message: Dict[str, str] = {}
		agency_to_message_workload: Dict[str, str] = {}
		agency_to_date_line: Dict[str, str] = {}
		error = None
		suggested_prefix = ""

		if did_fetch:
			selected_days = _parse_days(days_param)
			ordered_days = sorted(selected_days)

			# 날짜 매핑 생성 (YYYY-MM-DD 및 요일 포함 라벨)
			weekday_kr = ["월", "화", "수", "목", "금", "토", "일"]
			for d in ordered_days:
				cur = base_dt + timedelta(days=d)
				day_to_date[d] = cur.isoformat()
				day_to_date_label[d] = f"{cur.isoformat()}({weekday_kr[cur.weekday()]})"

			try:
				# 날짜별 그룹핑 (필터 모드 적용)
				grouped_by_date = fetch_grouped_messages_by_date(selected_days=selected_days, settings=settings, filter_mode=filter_mode)
			except Exception as e:
				grouped_by_date = {}
				error = str(e)
			else:
				error = None

			# 복붙 포맷 2종 생성
			# 1) 기본: 날짜(요일) → <작업명> → 상호명
			# 2) 작업량 포함: 날짜(요일) → <작업명> → 상호명 : 일작업량
			name_wl_re = re.compile(r"^(.+?)\s*\(일작업량\s+(.*?)\)$")
			for category, by_agency in grouped_by_date.items():
				for agency, by_day in by_agency.items():
					parts_base: List[str] = []
					parts_wl: List[str] = []
					for d in sorted(by_day.keys()):
						# 날짜 헤더
						date_label = day_to_date_label.get(d, f"+{d}")
						parts_base.append(date_label)
						parts_wl.append(date_label)
						# 작업명과 상호들
						for task, names in by_day[d].items():
							if not names:
								continue
							display_task = _strip_parentheses(task)
							parts_base.append(f"<{display_task}>")
							parts_wl.append(f"<{display_task}>")
							for name in names:
								name_str = str(name).strip()
								m = name_wl_re.match(name_str)
								if m:
									base_name = m.group(1).strip()
									workload_val = m.group(2).strip()
									parts_base.append(base_name)
									parts_wl.append(f"{base_name} : {workload_val}")
								else:
									# 작업량 정보가 없으면 동일하게 표기
									parts_base.append(name_str)
									parts_wl.append(name_str)
							# 작업 블록 사이: 1줄 공백
							parts_base.append("")
							parts_wl.append("")
						# 날짜 블록 사이: 추가로 1줄 더 공백(= 총 2줄)
						parts_base.append("")
						parts_wl.append("")
					agency_to_message[agency] = "\n".join(parts_base).rstrip()
					agency_to_message_workload[agency] = "\n".join(parts_wl).rstrip()

			# 대행사별 실제 존재하는 마감일 범위를 기준으로 날짜 문구(요일 포함) 생성
			weekday_kr = ["월", "화", "수", "목", "금", "토", "일"]
			def fmt_mmdd_w(dt: date) -> str:
				return f"{dt.month:02d}/{dt.day:02d}({weekday_kr[dt.weekday()]})"
			for category, by_agency in grouped_by_date.items():
				for agency, by_day in by_agency.items():
					present_days = sorted(list(by_day.keys()))
					if not present_days:
						continue
					start_dt = base_dt + timedelta(days=present_days[0])
					end_dt = base_dt + timedelta(days=present_days[-1])
					if start_dt == end_dt:
						line = f"{fmt_mmdd_w(start_dt)} 마감건 안내드립니다."
					else:
						line = f"{fmt_mmdd_w(start_dt)} ~ {fmt_mmdd_w(end_dt)} 마감건 안내드립니다."
					agency_to_date_line[agency] = line

			# 선택된 날짜 범위 기반 추천 첫 멘트 생성 (인사 + 날짜 문구, MM/DD 포맷)
			if ordered_days:
				all_dates = [base_dt + timedelta(days=d) for d in ordered_days]
				start_dt = min(all_dates)
				end_dt = max(all_dates)
				greeting = "대표님 안녕하세요~"
				def mmdd(dt: date) -> str:
					return f"{dt.month:02d}/{dt.day:02d}"
				if start_dt == end_dt:
					line = f"{mmdd(start_dt)} 마감건 안내드립니다."
				else:
					line = f"{mmdd(start_dt)} ~ {mmdd(end_dt)} 마감건 안내드립니다."
				suggested_prefix = greeting + "\n" + line

		# 마감일별 통계 계산 (0~5일)
		deadline_stats: Dict[int, List[str]] = {i: [] for i in range(6)}
		total_agency_count = 0
		if did_fetch and grouped_by_date:
			for category, by_agency in grouped_by_date.items():
				total_agency_count += len(by_agency)
				for agency, by_day in by_agency.items():
					for day in by_day.keys():
						if 0 <= day <= 5 and agency not in deadline_stats[day]:
							deadline_stats[day].append(agency)

		return render_template(
			"index.html",
			error=error,
			total_agency_count=total_agency_count,
			did_fetch=did_fetch,
			days_param=days_param,
			base_date_str=base_date_str or base_dt.isoformat(),
			day_to_date=day_to_date,
			day_to_date_label=day_to_date_label,
			ordered_days=ordered_days,
			grouped=grouped_by_date,
			agency_to_message=agency_to_message,
			agency_to_message_workload=agency_to_message_workload,
			agency_to_date_line=agency_to_date_line,
			deadline_stats=deadline_stats,
			settings=settings,
			filter_mode=filter_mode,
			suggested_prefix=suggested_prefix,
		)

	@app.route("/manage", methods=["GET"])  # 월보장 관리 대시보드
	@login_required
	def manage():
		return render_template("manage.html")

	@app.route("/settlement", methods=["GET"])  # 결재선 · 정산 페이지 (UI 스켈레톤)
	@login_required
	def settlement():
		from flask import send_file
		# 템플릿 엔진 경유 대신 파일을 직접 서빙하여, 템플릿 로더/캐시 이슈를 우회한다.
		return send_file(os.path.join(app.root_path, "templates", "settlement.html"), mimetype="text/html; charset=utf-8")

	# --- 결재선 보조 API들 ---
	@app.route("/api/settlement/tabs", methods=["GET"])  # 시트 탭 제목 목록 (결재선 전용 시트)
	def api_settlement_tabs():
		try:
			settlement_ssid = os.getenv("SETTLEMENT_SPREADSHEET_ID", "").strip()
			tabs = list_sheet_tabs(settlement_ssid or None)
		except Exception as e:
			return jsonify({"error": str(e)}), 500
		return jsonify({"tabs": tabs}), 200

	@app.route("/api/settlement/pricebook", methods=["GET", "POST"])  # 단가/계좌 저장소 - 파일 기반(로컬)
	def api_settlement_pricebook():
		storage_path = os.getenv("PRICEBOOK_PATH", os.path.join(os.getcwd(), "pricebook.json"))
		if request.method == "GET":
			try:
				if os.path.exists(storage_path):
					with open(storage_path, "r", encoding="utf-8") as f:
						data = json.load(f)
				else:
					data = []
			except Exception as e:
				return jsonify({"error": str(e)}), 500
			return jsonify({"items": data}), 200
		# POST: 저장
		try:
			payload = request.get_json(force=True, silent=False) or {}
		except Exception:
			return jsonify({"error": "invalid_json"}), 400
		items = payload.get("items")
		if not isinstance(items, list):
			return jsonify({"error": "invalid_items"}), 400
		try:
			with open(storage_path, "w", encoding="utf-8") as f:
				json.dump(items, f, ensure_ascii=False, indent=2)
		except Exception as e:
			return jsonify({"error": str(e)}), 500
		return jsonify({"ok": True}), 200

	@app.route("/api/settlement/pricebook/upload", methods=["POST"])  # XLSX 업로드 → 항목 파싱 반환
	def api_settlement_pricebook_upload():
		from io import BytesIO
		from openpyxl import load_workbook
		f = request.files.get("file")
		if not f:
			return jsonify({"error": "missing_file"}), 400
		try:
			buf = BytesIO(f.read())
			wb = load_workbook(buf, data_only=True)
			ws = wb.active
		except Exception as e:
			return jsonify({"error": f"xlsx_load_failed: {e}"}), 400
		# 헤더 매핑: 거래처, 상품명, 유형, 단가, 계좌, 예금주
		items = []
		try:
			headers = []
			for cell in ws[1]:
				headers.append(str(cell.value or "").strip())
			def idx(name: str) -> int | None:
				try:
					return headers.index(name)
				except ValueError:
					return None
			i_client = idx("거래처"); i_product = idx("상품명"); i_type = idx("유형"); i_price = idx("단가"); i_account = idx("계좌"); i_bank = idx("은행"); i_holder = idx("예금주")
			if i_client is None or i_product is None or i_price is None:
				return jsonify({"error": "missing_required_headers"}), 400
			for r in ws.iter_rows(min_row=2):
				def get(i):
					if i is None: return ""
					v = r[i].value if i < len(r) else ""
					return str(v).strip() if v is not None else ""
				client = get(i_client); product = get(i_product)
				if not client and not product:
					continue
				type_s = get(i_type)
				type_s = "공통" if type_s not in ("저장", "트래픽") else type_s
				price_s = get(i_price)
				try:
					price = float(str(price_s).replace(",", "")) if price_s else 0.0
				except Exception:
					price = 0.0
				account = get(i_account); bank = get(i_bank); holder = get(i_holder)
				items.append({"client": client, "product": product, "type": type_s, "price": price, "account": account, "bank": bank, "holder": holder})
		except Exception as e:
			return jsonify({"error": f"parse_failed: {e}"}), 400
		return jsonify({"items": items, "count": len(items)}), 200

	@app.route("/api/settlement/pricebook/template", methods=["GET"])  # 대량등록 XLSX 템플릿 다운로드
	def api_settlement_pricebook_template():
		from io import BytesIO
		from openpyxl import Workbook
		wb = Workbook()
		ws = wb.active
		ws.title = "pricebook"
		ws.append(["거래처", "상품명", "유형", "단가", "계좌", "은행", "예금주"])
		ws.append(["일류기획", "호올스", "저장", 32, "123-45-67890", "국민", "류준호"])  # 샘플
		buf = BytesIO()
		wb.save(buf)
		buf.seek(0)
		return app.response_class(buf.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=pricebook_template.xlsx"})

	@app.route("/api/settlement/inspect", methods=["GET"])  # 결재선 시트 헤더 점검
	def api_settlement_inspect():
		try:
			ssid = os.getenv("SETTLEMENT_SPREADSHEET_ID", "").strip()
			info = inspect_sheets_by_id(ssid)
		except Exception as e:
			return jsonify({"error": str(e)}), 500
		return jsonify({"tabs": info}), 200

	@app.route("/api/settlement/compute", methods=["POST"])  # 결재선 집계 계산
	def api_settlement_compute():
		try:
			payload = request.get_json(force=True, silent=False) or {}
		except Exception:
			return jsonify({"error": "invalid_json"}), 400
		ssid = os.getenv("SETTLEMENT_SPREADSHEET_ID", "").strip()
		selected_tabs = payload.get("tabs") or []
		if not isinstance(selected_tabs, list):
			return jsonify({"error": "invalid_tabs"}), 400
		# 단가 로드
		storage_path = os.getenv("PRICEBOOK_PATH", os.path.join(os.getcwd(), "pricebook.json"))
		try:
			if os.path.exists(storage_path):
				with open(storage_path, "r", encoding="utf-8") as f:
					pricebook = json.load(f)
			else:
				pricebook = []
		except Exception:
			pricebook = []
		try:
			result = compute_settlement_rows(ssid, selected_tabs, pricebook)
		except Exception as e:
			return jsonify({"error": str(e)}), 500
		return jsonify(result), 200

	@app.route("/api/settlement/extra", methods=["GET", "POST"])  # 수기 추가 지출 저장소
	def api_settlement_extra():
		storage_path = os.getenv("EXTRA_EXPENSES_PATH", os.path.join(os.getcwd(), "extra_expenses.json"))
		if request.method == "GET":
			try:
				if os.path.exists(storage_path):
					with open(storage_path, "r", encoding="utf-8") as f:
						data = json.load(f)
				else:
					data = []
			except Exception as e:
				return jsonify({"error": str(e)}), 500
			return jsonify({"items": data}), 200
		# POST 저장 (전체 치환 방식)
		try:
			payload = request.get_json(force=True, silent=False) or {}
		except Exception:
			return jsonify({"error": "invalid_json"}), 400
		items = payload.get("items")
		if not isinstance(items, list):
			return jsonify({"error": "invalid_items"}), 400
		try:
			with open(storage_path, "w", encoding="utf-8") as f:
				json.dump(items, f, ensure_ascii=False, indent=2)
		except Exception as e:
			return jsonify({"error": str(e)}), 500
		return jsonify({"ok": True}), 200

	@app.route("/api/settlement/cream2-accounts", methods=["GET", "POST"])  # 크림2 배포 계정 관리
	def api_settlement_cream2_accounts():
		# Render Disk 경로 우선 사용 (서버 재시작 시에도 데이터 유지)
		disk_path = "/var/data"
		default_path = os.path.join(disk_path, "cream2_accounts.json") if os.path.isdir(disk_path) else os.path.join(os.getcwd(), "cream2_accounts.json")
		storage_path = os.getenv("CREAM2_ACCOUNTS_PATH", default_path)
		if request.method == "GET":
			try:
				if os.path.exists(storage_path):
					with open(storage_path, "r", encoding="utf-8") as f:
						data = json.load(f)
				else:
					data = []
			except Exception as e:
				return jsonify({"error": str(e)}), 500
			return jsonify({"items": data}), 200
		# POST 저장 (전체 치환 방식)
		try:
			payload = request.get_json(force=True, silent=False) or {}
		except Exception:
			return jsonify({"error": "invalid_json"}), 400
		items = payload.get("items")
		if not isinstance(items, list):
			return jsonify({"error": "invalid_items"}), 400
		try:
			with open(storage_path, "w", encoding="utf-8") as f:
				json.dump(items, f, ensure_ascii=False, indent=2)
		except Exception as e:
			return jsonify({"error": str(e)}), 500
		return jsonify({"ok": True}), 200

	@app.route("/api/agency-pricing", methods=["GET", "POST"])  # 대행사 판매 단가 (우리가 받는 금액)
	def api_agency_pricing():
		# Render Disk 경로 우선 사용
		disk_path = "/var/data"
		default_path = os.path.join(disk_path, "agency_pricing.json") if os.path.isdir(disk_path) else os.path.join(os.getcwd(), "agency_pricing.json")
		storage_path = os.getenv("AGENCY_PRICING_PATH", default_path)
		if request.method == "GET":
			try:
				if os.path.exists(storage_path):
					with open(storage_path, "r", encoding="utf-8") as f:
						data = json.load(f)
				else:
					data = []
			except Exception as e:
				return jsonify({"error": str(e)}), 500
			return jsonify({"items": data}), 200
		# POST 저장
		try:
			payload = request.get_json(force=True, silent=False) or {}
		except Exception:
			return jsonify({"error": "invalid_json"}), 400
		items = payload.get("items")
		if not isinstance(items, list):
			return jsonify({"error": "invalid_items"}), 400
		try:
			with open(storage_path, "w", encoding="utf-8") as f:
				json.dump(items, f, ensure_ascii=False, indent=2)
		except Exception as e:
			return jsonify({"error": str(e)}), 500
		return jsonify({"ok": True}), 200

	@app.route("/debug/headers")
	def debug_headers():
		try:
			info = inspect_sheets()
		except Exception as e:
			return jsonify({"error": str(e)}), 500
		return jsonify(info)

	@app.route("/debug/matches")
	def debug_matches():
		days_param = request.args.get("days", "0").strip()
		selected_days = _parse_days(days_param)
		try:
			report = diagnose_matches(selected_days)
		except Exception as e:
			return jsonify({"error": str(e)}), 500
		return jsonify(report)

	@app.route("/api/fetch-stream")
	def fetch_stream():
		"""SSE: 진행률과 최종 결과를 스트리밍한다."""
		settings = load_settings()
		days_param = request.args.get("days", "").strip()
		filter_mode = request.args.get("filter_mode", "agency").strip().lower()
		selected_days = _parse_days(days_param)

		def event_stream():
			try:
				for evt in stream_grouped_messages_by_date(selected_days, settings, filter_mode):
					yield f"data: {json.dumps(evt, ensure_ascii=False)}\n\n"
			except Exception as e:
				payload = {"type": "error", "message": str(e)}
				yield f"data: {json.dumps(payload, ensure_ascii=False)}\n\n"

		return Response(stream_with_context(event_stream()), mimetype="text/event-stream")

	@app.route("/api/internal/items", methods=["GET"])  # 캐시된 내부 진행건 목록 반환
	def api_internal_items():
		try:
			cache = internal_load_cache()
		except Exception as e:
			return jsonify({"error": str(e)}), 500
		return jsonify({
			"updated_at": cache.get("updated_at"),
			"items": cache.get("items", []),
		}), 200

	@app.route("/api/internal/refresh", methods=["POST"])  # 수동 불러오기
	def api_internal_refresh():
		try:
			data = internal_refresh_cache()
		except Exception as e:
			return jsonify({"error": str(e)}), 500
		return jsonify(data), 200
	
	@app.route("/api/workload/schedule", methods=["GET"])  # 작업량 스케줄 조회
	def api_workload_schedule():
		"""최근 3주간 작업량 스케줄 조회 (캐시 우선)"""
		company = request.args.get("company")
		business_name = request.args.get("business_name")  # 업체 필터 추가
		
		try:
			schedule = fetch_workload_schedule(company, business_name)
			return jsonify(schedule), 200
		except Exception as e:
			logger.error(f"Workload schedule error: {e}")
			import traceback
			logger.error(traceback.format_exc())
			return jsonify({"error": str(e)}), 500
	
	@app.route("/api/workload/cache/refresh", methods=["POST"])  # 작업량 캐시 수동 갱신
	def api_workload_cache_refresh():
		"""작업량 캐시 수동 갱신"""
		try:
			from workload_cache import refresh_all_workload_cache
			result = refresh_all_workload_cache()
			return jsonify(result), 200
		except Exception as e:
			logger.error(f"Workload cache refresh error: {e}")
			import traceback
			logger.error(traceback.format_exc())
			return jsonify({"error": str(e)}), 500
	
	@app.route("/api/workload/cache/status", methods=["GET"])  # 작업량 캐시 상태 조회
	def api_workload_cache_status():
		"""작업량 캐시 상태 조회"""
		try:
			from workload_cache import WorkloadCache
			cache = WorkloadCache()
			status = cache.get_cache_status()
			return jsonify(status), 200
		except Exception as e:
			logger.error(f"Workload cache status error: {e}")
			return jsonify({"error": str(e)}), 500
	
	@app.route("/api/workload/businesses", methods=["GET"])  # 업체별 작업량 일괄 조회
	def api_workload_businesses():
		"""특정 회사의 모든 업체별 작업량 데이터 조회 (캐시 우선)"""
		company = request.args.get("company")
		
		if not company:
			return jsonify({"error": "company parameter required"}), 400
		
		try:
			from workload_cache import WorkloadCache
			cache = WorkloadCache()
			
			# 캐시에서 업체별 데이터 조회
			businesses_data = cache.get_all_businesses_workload(company)
			
			if businesses_data:
				logger.info(f"Loaded {len(businesses_data)} businesses from cache for {company}")
				return jsonify({
					"businesses": businesses_data,
					"from_cache": True,
					"count": len(businesses_data)
				}), 200
			else:
				logger.warning(f"No cached data for {company}, cache may be expired")
				return jsonify({
					"businesses": {},
					"from_cache": False,
					"count": 0,
					"message": "캐시가 없거나 만료되었습니다. '작업량 갱신' 버튼을 눌러주세요."
				}), 200
		except Exception as e:
			logger.error(f"Business workload error: {e}")
			import traceback
			logger.error(traceback.format_exc())
			return jsonify({"error": str(e)}), 500

	@app.route("/api/mark-done", methods=["POST"])
	def mark_done():
		"""특정 대행사 카드의 모든 해당 행을 '마감 안내 체크'로 표시한다."""
		try:
			data = request.get_json(force=True, silent=False) or {}
		except Exception:
			return jsonify({"error": "invalid_json"}), 400

		agency_label = str(data.get("agency") or "").strip()
		days_param = str(data.get("days") or "").strip()
		filter_mode = str(data.get("filter_mode") or "agency").strip().lower()
		if not agency_label:
			return jsonify({"error": "missing_agency"}), 400

		selected_days = _parse_days(days_param)
		try:
			result = mark_checked_for_agency(selected_days=selected_days, agency_label=agency_label, filter_mode=filter_mode)
		except Exception as e:
			return jsonify({"error": str(e)}), 500
		return jsonify(result)

	@app.route("/api/mark-done-bulk", methods=["POST"])
	def mark_done_bulk():
		"""여러 대행사 카드를 한 번에 체크 처리한다."""
		try:
			data = request.get_json(force=True, silent=False) or {}
		except Exception:
			return jsonify({"error": "invalid_json"}), 400

		agency_labels = data.get("agencies") or []
		if not isinstance(agency_labels, list):
			return jsonify({"error": "invalid_agencies"}), 400
		agency_labels = [str(a or "").strip() for a in agency_labels if str(a or "").strip()]
		if not agency_labels:
			return jsonify({"error": "empty_agencies"}), 400

		days_param = str(data.get("days") or "").strip()
		filter_mode = str(data.get("filter_mode") or "agency").strip().lower()
		selected_days = _parse_days(days_param)
		try:
			result = mark_checked_for_agencies(selected_days=selected_days, agency_labels=agency_labels, filter_mode=filter_mode)
		except Exception as e:
			return jsonify({"error": str(e)}), 500
		return jsonify(result)

	# --- 월보장 관리 API ---
	@app.route("/api/guarantee/items", methods=["GET", "POST"])
	def api_guarantee_items():
		"""보장건 목록 조회 및 생성"""
		gm = GuaranteeManager()
		
		if request.method == "GET":
			# 필터 파라미터
			filters = {}
			if request.args.get("company"):
				filters["company"] = request.args.get("company")
			if request.args.get("status"):
				filters["status"] = request.args.get("status")
			if request.args.get("product"):
				filters["product"] = request.args.get("product")
			if request.args.get("active_only"):
				filters["active_only"] = True
			
			logger.info(f"Getting items with filters: {filters}")
			items = gm.get_items(filters)
			
			# 데이터가 없으면 자동 동기화 시도 (서버 재시작 후 첫 요청)
			if len(gm.get_items()) == 0:
				logger.info("📦 No local data found. Auto-syncing from Google Sheets...")
				try:
					sync_result = gm.sync_from_google_sheets()
					logger.info(f"✅ Auto-sync completed: Added {sync_result.get('added', 0)} items")
					# 다시 데이터 조회
					items = gm.get_items(filters)
				except Exception as sync_err:
					logger.error(f"❌ Auto-sync failed: {sync_err}")
			
			logger.info(f"Found {len(items)} items")
			
			# 디버깅: 처음 몇 개 아이템 로그
			if items:
				logger.info(f"Sample item: {items[0] if items else 'No items'}")
			
			return jsonify({"items": items, "count": len(items)}), 200
		
		# POST: 새 보장건 생성
		try:
			data = request.get_json(force=True)
		except Exception:
			return jsonify({"error": "invalid_json"}), 400
		
		if not data.get("business_name"):
			return jsonify({"error": "business_name_required"}), 400
		
		item = gm.create_item(data)
		return jsonify(item), 201

	@app.route("/api/guarantee/items/<item_id>", methods=["GET", "PUT", "DELETE"])
	def api_guarantee_item(item_id):
		"""특정 보장건 조회/수정/삭제"""
		gm = GuaranteeManager()
		
		if request.method == "GET":
			item = gm.get_item(item_id)
			if not item:
				return jsonify({"error": "not_found"}), 404
			return jsonify(item), 200
		
		elif request.method == "PUT":
			try:
				data = request.get_json(force=True)
			except Exception:
				return jsonify({"error": "invalid_json"}), 400
			
			item = gm.update_item(item_id, data)
			if not item:
				return jsonify({"error": "not_found"}), 404
			return jsonify(item), 200
		
		elif request.method == "DELETE":
			if gm.delete_item(item_id):
				return jsonify({"ok": True}), 200
			return jsonify({"error": "not_found"}), 404

	@app.route("/api/guarantee/statistics", methods=["GET"])
	def api_guarantee_stats():
		"""통계 조회"""
		gm = GuaranteeManager()
		return jsonify(gm.get_statistics()), 200

	@app.route("/api/guarantee/search", methods=["GET"])
	def api_guarantee_search():
		"""검색"""
		query = request.args.get("q", "").strip()
		if not query:
			return jsonify({"items": []}), 200
		
		gm = GuaranteeManager()
		items = gm.search(query)
		return jsonify({"items": items, "count": len(items)}), 200

	@app.route("/api/guarantee/daily-rank", methods=["POST"])
	def api_guarantee_daily_rank():
		"""일차별 순위 업데이트"""
		try:
			data = request.get_json(force=True)
		except Exception:
			return jsonify({"error": "invalid_json"}), 400
		
		item_id = data.get("item_id")
		day = data.get("day")
		rank = data.get("rank")
		
		if not all([item_id, day is not None, rank is not None]):
			return jsonify({"error": "missing_params"}), 400
		
		gm = GuaranteeManager()
		item = gm.update_daily_rank(item_id, day, rank)
		if not item:
			return jsonify({"error": "not_found"}), 404
		return jsonify(item), 200

	@app.route("/api/guarantee/sync", methods=["POST"])
	def api_guarantee_sync():
		"""수동 동기화 - 단계별 진행 (UI에서 진행 상태 표시용)
		
		플로우:
		1. 제이투랩 업체 정보 가져오기
		2. 일류기획 업체 정보 가져오기
		3. 오늘 순위 데이터 확인 (rank_snapshots)
		4. 시간 체크 후 크롤링 (00:00~15:10 사이면 스킵)
		5. 시트에 순위 기입
		6. 작업량 데이터 갱신
		"""
		import pytz
		from datetime import datetime
		
		kst = pytz.timezone('Asia/Seoul')
		now = datetime.now(kst)
		current_hour = now.hour
		current_minute = now.minute
		today_str = now.strftime("%Y-%m-%d")
		
		# 결과 저장용
		steps = {}
		
		try:
			# ============ STEP 1 & 2: 업체 정보 가져오기 ============
			gm = GuaranteeManager()
			logger.info("📡 Starting sync: fetching company data...")
			
			sync_result = gm.sync_from_google_sheets()
			
			# 회사별 카운트
			jtwolab_items = len(gm.get_items({"company": "제이투랩"}))
			ilryu_items = len(gm.get_items({"company": "일류기획"}))
			total_items = jtwolab_items + ilryu_items
			
			steps["jtwolab_sync"] = {
				"status": "success",
				"count": jtwolab_items,
				"message": f"제이투랩 {jtwolab_items}건"
			}
			steps["ilryu_sync"] = {
				"status": "success", 
				"count": ilryu_items,
				"message": f"일류기획 {ilryu_items}건"
			}
			
			logger.info(f"✅ Company data fetched: 제이투랩 {jtwolab_items}, 일류기획 {ilryu_items}")
			
			# ============ STEP 3: 오늘 순위 데이터 확인 ============
			has_today_rank = False
			try:
				from rank_snapshot_manager import RankSnapshotManager
				rsm = RankSnapshotManager()
				today_snapshots = rsm.get_history(date_from=today_str, date_to=today_str, days=1)
				has_today_rank = bool(today_snapshots and len(today_snapshots) > 0)
				
				steps["rank_check"] = {
					"status": "success",
					"has_data": has_today_rank,
					"count": len(today_snapshots) if today_snapshots else 0,
					"message": f"오늘 순위 {'있음' if has_today_rank else '없음'}"
				}
				logger.info(f"🔍 Rank check: {'데이터 있음' if has_today_rank else '데이터 없음'} ({len(today_snapshots) if today_snapshots else 0}건)")
			except Exception as e:
				steps["rank_check"] = {"status": "error", "message": str(e)}
				logger.warning(f"Rank check failed: {e}")
			
			# ============ STEP 4: 시간 체크 후 크롤링 ============
			# 00:00~15:09 사이면 크롤링 스킵
			is_crawl_time_window = (current_hour >= 0 and current_hour < 15) or (current_hour == 15 and current_minute < 10)
			
			rank_crawled = False
			sheet_updated = False
			
			if has_today_rank:
				# 이미 오늘 데이터가 있으면 크롤링 스킵
				steps["rank_crawl"] = {
					"status": "skipped",
					"reason": "already_exists",
					"message": "오늘 데이터 이미 존재"
				}
				logger.info("🔄 Today's rank data exists, skipping crawl but will try sheet update")
				
				# 시트 업데이트는 시도 (시트에 아직 안 기입됐을 수 있음)
				try:
					from rank_update_service import update_guarantee_sheets_from_snapshots
					from scheduler_logs import log_scheduler_event
					log_scheduler_event("guarantee_update", "보장건 시트 업데이트", "started", "기존 데이터로 시트 업데이트")
					logger.info("📝 Updating guarantee sheets with existing data...")
					
					update_result = update_guarantee_sheets_from_snapshots()
					total_updated = update_result.get('total_updated', 0)
					
					steps["sheet_update"] = {
						"status": "success",
						"count": total_updated,
						"message": f"{total_updated}건 시트 기입"
					}
					sheet_updated = True
					log_scheduler_event("guarantee_update", "보장건 시트 업데이트", "success", f"{total_updated}건 업데이트")
					logger.info(f"✅ Sheet update completed: {total_updated}건")
				except Exception as ue:
					steps["sheet_update"] = {"status": "error", "message": str(ue)}
					log_scheduler_event("guarantee_update", "보장건 시트 업데이트", "failed", str(ue))
					logger.error(f"❌ Sheet update failed: {ue}")
					
			elif is_crawl_time_window:
				# 00:00~15:09 사이면 스킵
				steps["rank_crawl"] = {
					"status": "skipped",
					"reason": "time_window",
					"message": f"00:00~15:10 사이 ({now.strftime('%H:%M')})"
				}
				steps["sheet_update"] = {
					"status": "skipped",
					"reason": "time_window",
					"message": "크롤링 스킵됨"
				}
				logger.info(f"⏰ Crawl skipped: time window (current: {now.strftime('%H:%M')})")
			else:
				# 크롤링 실행
				try:
					from rank_crawler import crawl_ranks_for_company
					from scheduler_logs import log_scheduler_event
					
					log_scheduler_event("rank_crawl", "순위 크롤링 (동기화)", "started", "동기화 버튼으로 실행")
					logger.info("🏆 Starting rank crawl...")
					
					crawl_result = crawl_ranks_for_company(None)
					crawled_count = crawl_result.get('crawled_count', 0)
					
					steps["rank_crawl"] = {
						"status": "success",
						"count": crawled_count,
						"message": f"{crawled_count}건 크롤링 완료"
					}
					rank_crawled = True
					log_scheduler_event("rank_crawl", "순위 크롤링 (동기화)", "success", f"{crawled_count}건 완료")
					logger.info(f"✅ Rank crawl completed: {crawled_count}건")
					
					# ============ STEP 5: 시트에 순위 기입 ============
					try:
						from rank_update_service import update_guarantee_sheets_from_snapshots
						log_scheduler_event("guarantee_update", "보장건 시트 업데이트", "started", "동기화 후 실행")
						logger.info("📝 Updating guarantee sheets...")
						
						update_result = update_guarantee_sheets_from_snapshots()
						total_updated = update_result.get('total_updated', 0)
						
						steps["sheet_update"] = {
							"status": "success",
							"count": total_updated,
							"message": f"{total_updated}건 시트 기입"
						}
						sheet_updated = True
						log_scheduler_event("guarantee_update", "보장건 시트 업데이트", "success", f"{total_updated}건 업데이트")
						logger.info(f"✅ Sheet update completed: {total_updated}건")
					except Exception as ue:
						steps["sheet_update"] = {"status": "error", "message": str(ue)}
						log_scheduler_event("guarantee_update", "보장건 시트 업데이트", "failed", str(ue))
						logger.error(f"❌ Sheet update failed: {ue}")
						
				except Exception as ce:
					steps["rank_crawl"] = {"status": "error", "message": str(ce)}
					steps["sheet_update"] = {"status": "skipped", "reason": "crawl_failed"}
					from scheduler_logs import log_scheduler_event
					log_scheduler_event("rank_crawl", "순위 크롤링 (동기화)", "failed", str(ce))
					logger.error(f"❌ Rank crawl failed: {ce}")
			
			# ============ STEP 6: 작업량 데이터 갱신 ============
			workload_refreshed = False
			try:
				from workload_cache import WorkloadCache, refresh_all_workload_cache
				wc = WorkloadCache()
				
				if not wc.is_cache_valid():
					logger.info("⚡ Refreshing workload cache...")
					wresult = refresh_all_workload_cache()
					workload_refreshed = True
					steps["workload_refresh"] = {
						"status": "success",
						"message": "작업량 캐시 갱신 완료"
					}
					logger.info(f"✅ Workload cache refreshed")
				else:
					steps["workload_refresh"] = {
						"status": "skipped",
						"reason": "cache_valid",
						"message": "캐시 유효 (갱신 불필요)"
					}
			except Exception as we:
				steps["workload_refresh"] = {"status": "error", "message": str(we)}
				logger.warning(f"Workload refresh failed: {we}")
			
			# ============ 최종 결과 ============
			last_sync = gm.get_last_sync_time()
			
			# 메시지 생성
			message_parts = [f"동기화 완료 (총 {total_items}건)"]
			if rank_crawled:
				message_parts.append("순위 크롤링됨")
			if sheet_updated:
				message_parts.append("시트 기입됨")
			if workload_refreshed:
				message_parts.append("작업량 갱신됨")
			
			return jsonify({
				"ok": True,
				"steps": steps,
				"result": sync_result,
				"last_sync": last_sync,
				"total_items": total_items,
				"jtwolab_count": jtwolab_items,
				"ilryu_count": ilryu_items,
				"rank_crawled": rank_crawled,
				"sheet_updated": sheet_updated,
				"workload_refreshed": workload_refreshed,
				"message": " · ".join(message_parts)
			}), 200
			
		except Exception as e:
			logger.error(f"Manual sync failed: {str(e)}")
			import traceback
			logger.error(f"Traceback: {traceback.format_exc()}")
			return jsonify({
				"ok": False,
				"error": str(e),
				"steps": steps,
				"detail": "서버 로그를 확인하세요"
			}), 500


	@app.route("/api/guarantee/sync-status", methods=["GET"])
	def api_guarantee_sync_status():
		"""동기화 상태 확인"""
		gm = GuaranteeManager()
		last_sync = gm.get_last_sync_time()
		
		# 다음 동기화 시간 계산
		now = datetime.now(pytz.timezone('Asia/Seoul'))
		current_hour = now.hour
		
		if current_hour < 9:
			next_sync = now.replace(hour=9, minute=0, second=0, microsecond=0)
		elif current_hour < 16:
			next_sync = now.replace(hour=16, minute=0, second=0, microsecond=0)
		else:
			# 다음날 9시
			next_sync = (now + timedelta(days=1)).replace(hour=9, minute=0, second=0, microsecond=0)
		
		return jsonify({
			"last_sync": last_sync,
			"next_sync": next_sync.isoformat(),
			"next_sync_kst": next_sync.strftime("%Y-%m-%d %H:%M KST")
		}), 200

	@app.route("/api/guarantee/exposure-status", methods=["GET"])
	def api_exposure_status():
		"""실시간 노출 현황 조회"""
		company = request.args.get("company")  # 제이투랩, 일류기획
		
		try:
			gm = GuaranteeManager()
			status = gm.get_exposure_status(company)
			
			# 크롤링 순위 데이터 병합
			try:
				from rank_crawler import get_latest_ranks
				latest_ranks = get_latest_ranks(company)
				
				# 상호명 -> 순위 매핑
				rank_map = {}
				for rank_data in latest_ranks:
					rank_map[rank_data["business_name"]] = {
						"rank": rank_data["rank"],
						"keyword": rank_data["keyword"],
						"checked_at": rank_data["checked_at"]
					}
				
				# exposure_details에 크롤링 순위 추가
				for detail in status.get("exposure_details", []):
					biz_name = detail.get("business_name")
					if biz_name in rank_map:
						detail["crawled_rank"] = rank_map[biz_name]["rank"]
						detail["crawled_at"] = rank_map[biz_name]["checked_at"]
			except Exception as e:
				logger.warning(f"Failed to merge crawled ranks: {e}")
			
			return jsonify(status), 200
		except Exception as e:
			logger.error(f"Exposure status error: {e}")
			return jsonify({"error": str(e)}), 500
	
	@app.route("/api/guarantee/deadline-status", methods=["GET"])
	def api_deadline_status():
		"""마감 임박 현황 조회"""
		company = request.args.get("company")  # 제이투랩, 일류기획
		
		try:
			gm = GuaranteeManager()
			status = gm.get_deadline_status(company)
			return jsonify(status), 200
		except Exception as e:
			logger.error(f"Deadline status error: {e}")
			return jsonify({"error": str(e)}), 500
	
	@app.route("/api/recipe-analysis", methods=["GET"])
	def api_recipe_analysis():
		"""최적 레시피 분석 결과 조회"""
		weeks = request.args.get("weeks", 3, type=int)
		weeks = min(max(weeks, 1), 3)  # 1~3주 제한
		
		try:
			from recipe_analyzer import get_analyzer
			analyzer = get_analyzer()
			result = analyzer.analyze_all(weeks=weeks)
			return jsonify(result), 200
		except Exception as e:
			logger.error(f"Recipe analysis error: {e}")
			import traceback
			logger.error(traceback.format_exc())
			return jsonify({"error": str(e)}), 500
	
	@app.route("/api/business-dashboard/<business_name>", methods=["GET"])
	def api_business_dashboard(business_name):
		"""업체별 대시보드 데이터"""
		try:
			from recipe_analyzer import get_analyzer
			analyzer = get_analyzer()
			result = analyzer.get_business_dashboard(business_name)
			
			if result:
				return jsonify(result), 200
			else:
				return jsonify({"error": "Business not found"}), 404
		except Exception as e:
			logger.error(f"Business dashboard error: {e}")
			return jsonify({"error": str(e)}), 500
	
	@app.route("/api/guarantee/security-status", methods=["GET"])
	def api_security_status():
		"""데이터 보안 상태 확인 (관리자용)"""
		status = {
			"encryption_enabled": SECURITY_AVAILABLE,
			"data_info": {}
		}
		
		if SECURITY_AVAILABLE:
			try:
				security = DataSecurity()
				status["data_info"] = security.get_data_info()
				status["encryption_key_source"] = "environment" if os.getenv("DATA_ENCRYPTION_KEY") else "file"
			except Exception as e:
				status["error"] = str(e)
		else:
			status["warning"] = "Encryption module not available"
		
		# 기본 데이터 파일 존재 여부
		status["plain_files"] = {
			"guarantee_data.json": os.path.exists("guarantee_data.json"),
			"pricebook.json": os.path.exists("pricebook.json"),
			"internal_cache.json": os.path.exists("internal_cache.json")
		}
		
		return jsonify(status), 200
	
	@app.route("/api/guarantee/export", methods=["GET"])
	def api_guarantee_export():
		"""보장건 데이터 내보내기 (JSON)"""
		try:
			gm = GuaranteeManager()
			items = gm.get_items()
			
			# 민감 정보 제거 옵션
			remove_sensitive = request.args.get("remove_sensitive", "false").lower() == "true"
			if remove_sensitive:
				for item in items:
					item.pop("place_account", None)
					item.pop("url", None)
			
			return jsonify({
				"exported_at": datetime.now().isoformat(),
				"count": len(items),
				"items": items
			}), 200
		except Exception as e:
			return jsonify({"error": str(e)}), 500
	
	@app.route("/api/guarantee/crawl-ranks", methods=["POST"])
	def api_crawl_ranks():
		"""애드로그에서 순위 크롤링 실행 (N2 포함, Google Sheets 저장)"""
		company = request.args.get("company")  # 제이투랩, 일류기획 (None이면 전체)
		
		try:
			from rank_crawler import crawl_ranks_for_company
			result = crawl_ranks_for_company(company)
			
			if result.get("success"):
				# 보장건 시트 자동 업데이트 (성공 시)
				try:
					from rank_update_service import update_guarantee_sheets_from_snapshots
					update_result = update_guarantee_sheets_from_snapshots()
					result["sheet_update"] = update_result
					logger.info(f"✅ Guarantee sheets updated manually: {update_result}")
				except Exception as update_error:
					logger.error(f"❌ Guarantee sheet update failed: {update_error}")
					result["sheet_update_error"] = str(update_error)
					
				return jsonify(result), 200
			else:
				return jsonify(result), 500
		except ValueError as ve:
			# 환경변수 오류 (ADLOG_ID/PASSWORD 미설정)
			logger.error(f"Configuration error: {ve}")
			return jsonify({"success": False, "error": str(ve), "message": "환경변수 오류"}), 500
		except Exception as e:
			logger.error(f"Rank crawling error: {e}")
			import traceback
			logger.error(traceback.format_exc())
			return jsonify({"success": False, "error": str(e), "message": "크롤링 실패"}), 500
	
	@app.route("/api/cron/crawl-ranks", methods=["POST"])
	def api_cron_crawl_ranks():
		"""외부 cron 서비스용 순위 크롤링 endpoint (token 인증)
		
		Render 중복 실행 방지를 위해 외부 cron 서비스(cron-job.org 등)를 사용할 때:
		- CRON_TOKEN 환경변수 설정 필수
		- Header: Authorization: Bearer <token> 또는 Query: ?token=<token>
		"""
		# 토큰 추출
		auth_header = request.headers.get("Authorization", "")
		if auth_header.startswith("Bearer "):
			token = auth_header[7:]
		else:
			token = request.args.get("token") or request.form.get("token")
		
		# 토큰 검증
		expected_token = os.getenv("CRON_TOKEN")
		if not expected_token:
			logger.warning("CRON_TOKEN not configured")
			return jsonify({"success": False, "message": "CRON_TOKEN 미설정"}), 500
		
		if not token or token != expected_token:
			logger.warning(f"Invalid cron token attempt")
			return jsonify({"success": False, "message": "Invalid token"}), 401
		
		# 크롤링 실행
		try:
			from rank_crawler import crawl_ranks_for_company
			result = crawl_ranks_for_company(None)  # 전체 회사
			
			if result.get("success"):
				logger.info(f"✅ Cron crawl completed: {result.get('message')}")
				return jsonify(result), 200
			else:
				return jsonify(result), 500
		except Exception as e:
			logger.error(f"Cron crawl error: {e}")
			return jsonify({"success": False, "error": str(e)}), 500
	
	@app.route("/api/guarantee/latest-ranks", methods=["GET"])
	def api_latest_ranks():
		"""최신 순위 데이터 조회"""
		company = request.args.get("company")
		
		try:
			from rank_crawler import get_latest_ranks
			ranks = get_latest_ranks(company)
			return jsonify({"ranks": ranks, "count": len(ranks)}), 200
		except Exception as e:
			logger.error(f"Latest ranks error: {e}")
			return jsonify({"error": str(e)}), 500
	
	@app.route("/api/guarantee/rank-history/<business_name>", methods=["GET"])
	def api_rank_history(business_name):
		"""업체별 순위 히스토리 조회"""
		limit = int(request.args.get("limit", 30))
		
		try:
			from rank_crawler import get_rank_history
			history = get_rank_history(business_name, limit)
			return jsonify({"history": history, "count": len(history)}), 200
		except Exception as e:
			logger.error(f"Rank history error: {e}")
			return jsonify({"error": str(e)}), 500
	
	@app.route("/api/guarantee/ranks/export", methods=["GET"])
	def api_ranks_export():
		"""순위 데이터 JSON으로 내보내기 (백업용)"""
		try:
			from db_backup import export_rank_history_to_json
			data = export_rank_history_to_json()
			return jsonify(data), 200
		except Exception as e:
			logger.error(f"Rank export error: {e}")
			return jsonify({"error": str(e)}), 500
	
	@app.route("/api/guarantee/ranks/import", methods=["POST"])
	def api_ranks_import():
		"""순위 데이터 JSON에서 가져오기 (복원용)"""
		try:
			data = request.get_json(force=True)
			from db_backup import import_rank_history_from_json
			success = import_rank_history_from_json(data)
			
			if success:
				return jsonify({"ok": True, "message": "순위 데이터 복원 완료"}), 200
			else:
				return jsonify({"error": "복원 실패"}), 500
		except Exception as e:
			logger.error(f"Rank import error: {e}")
			return jsonify({"error": str(e)}), 500

	# --- Worklog Cache API ---
	@app.route("/api/worklog/cache/refresh", methods=["POST"])
	@login_required
	def api_worklog_cache_refresh():
		"""Worklog 캐시 갱신"""
		try:
			from worklog_cache import refresh_worklog_cache
			result = refresh_worklog_cache()
			return jsonify(result), 200 if result.get("success") else 500
		except Exception as e:
			logger.error(f"Worklog cache refresh error: {e}")
			return jsonify({"error": str(e)}), 500
	
	@app.route("/api/worklog/cache/status", methods=["GET"])
	@login_required
	def api_worklog_cache_status():
		"""Worklog 캐시 상태 조회"""
		try:
			from worklog_cache import get_worklog_cache_status
			status = get_worklog_cache_status()
			return jsonify(status), 200
		except Exception as e:
			logger.error(f"Worklog cache status error: {e}")
			return jsonify({"error": str(e)}), 500
	
	# --- Training Dataset API ---
	@app.route("/api/training/build", methods=["POST"])
	@login_required
	def api_training_build():
		"""학습 데이터셋 빌드"""
		weeks = request.args.get("weeks", 3, type=int)
		weeks = min(max(weeks, 1), 8)  # 1~8주 제한
		
		try:
			from training_dataset_builder import build_and_save
			result = build_and_save(weeks=weeks)
			return jsonify(result), 200 if result.get("success") else 500
		except Exception as e:
			logger.error(f"Training build error: {e}")
			import traceback
			logger.error(traceback.format_exc())
			return jsonify({"error": str(e)}), 500
	
	@app.route("/api/recipe/top", methods=["GET"])
	@login_required
	def api_recipe_top():
		"""상위 레시피 조회"""
		weeks = request.args.get("weeks", 3, type=int)
		weeks = min(max(weeks, 1), 8)
		
		try:
			from training_dataset_builder import get_top_recipes
			recipes = get_top_recipes(weeks=weeks)
			return jsonify({"recipes": recipes, "count": len(recipes)}), 200
		except Exception as e:
			logger.error(f"Recipe top error: {e}")
			return jsonify({"error": str(e)}), 500

	# --- Scheduler Logs API ---
	@app.route("/api/scheduler/logs", methods=["GET"])
	@login_required
	def api_scheduler_logs():
		"""스케줄러 로그 조회"""
		job_id = request.args.get("job_id")
		status = request.args.get("status")
		limit = request.args.get("limit", 50, type=int)
		limit = min(max(limit, 1), 100)
		
		try:
			from scheduler_logs import get_scheduler_logs
			logs = get_scheduler_logs(job_id=job_id, status=status, limit=limit)
			return jsonify({"logs": logs, "count": len(logs)}), 200
		except Exception as e:
			logger.error(f"Scheduler logs error: {e}")
			return jsonify({"error": str(e)}), 500
	
	@app.route("/api/scheduler/summary", methods=["GET"])
	@login_required
	def api_scheduler_summary():
		"""스케줄러 요약 조회"""
		try:
			from scheduler_logs import get_scheduler_summary
			summary = get_scheduler_summary()
			return jsonify(summary), 200
		except Exception as e:
			logger.error(f"Scheduler summary error: {e}")
			return jsonify({"error": str(e)}), 500

	# Shutdown scheduler when app closes
	import atexit
	atexit.register(lambda: scheduler.shutdown())

	return app

app = create_app()


def _parse_days(days_param: str) -> List[int]:
	if not days_param:
		return []
	parts = [p.strip() for p in days_param.split(",") if p.strip()]
	selected: List[int] = []
	for p in parts:
		try:
			selected.append(int(p))
		except ValueError:
			continue
	return selected


if __name__ == "__main__":
	parser = argparse.ArgumentParser()
	parser.add_argument("--prod", action="store_true", help="Run with waitress server")
	args = parser.parse_args()

	host = os.getenv("FLASK_HOST", "0.0.0.0")
	port = int(os.getenv("FLASK_PORT", "8080"))
	debug = os.getenv("FLASK_DEBUG", "false").lower() == "true"

	if args.prod:
		from waitress import serve

		serve(app, host=host, port=port)
	else:
		app.run(host=host, port=port, debug=debug)
